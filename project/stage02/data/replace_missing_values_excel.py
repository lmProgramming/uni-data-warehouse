import openpyxl
import os


def replace_text_in_excel(filepath, text_to_find, text_to_replace_with, case_sensitive=False, backup=True):
    """
    Replaces text in all string cells across all sheets of an Excel file.

    Args:
        filepath (str): The path to the Excel file.
        text_to_find (str): The text string to search for.
        text_to_replace_with (str): The text string to replace with.
        case_sensitive (bool): Whether the search should be case-sensitive.
        backup (bool): If True, creates a backup of the original file.
    """
    if not os.path.exists(filepath):
        print(f"Error: File not found at {filepath}")
        return

    if backup:
        base, ext = os.path.splitext(filepath)
        backup_filepath = f"{base}_backup{ext}"
        try:
            import shutil
            shutil.copyfile(filepath, backup_filepath)
            print(f"Backup created: {backup_filepath}")
        except Exception as e:
            print(f"Could not create backup: {e}")
            # Optionally, you could decide to not proceed if backup fails
            # return

    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(filepath)
        print(f"Processing workbook: {filepath}")
        total_replacements_in_workbook = 0

        # Iterate through all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"  Processing sheet: {sheet.title}...")
            replacements_in_sheet = 0

            # Iterate through all rows and then all cells in each row
            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    # Check if the cell value is a string
                    if cell.value is not None and isinstance(cell.value, str):
                        original_value = cell.value
                        new_value = ""

                        if case_sensitive:
                            if text_to_find in original_value:
                                new_value = original_value.replace(
                                    text_to_find, text_to_replace_with)
                        else:
                            # For case-insensitive, we need a bit more work or regex
                            # Simple approach: convert both to lower for comparison, then replace original
                            if text_to_find.lower() in original_value.lower():
                                # This simple replace might not be perfect for all case-insensitive scenarios
                                # if multiple mixed-case versions exist and text_to_find is short.
                                # A regex replace would be more robust for complex case-insensitivity.
                                # For "/N" vs "/n", this should be fine:
                                temp_val = original_value
                                # Replace both common cases if not case sensitive
                                if "/N" in temp_val:
                                    temp_val = temp_val.replace(
                                        "/N", text_to_replace_with)
                                if "/n" in temp_val:  # if "/N" was already replaced, this still works
                                    temp_val = temp_val.replace(
                                        "/n", text_to_replace_with)
                                new_value = temp_val

                        if new_value != original_value and new_value != "":  # Check if a replacement actually happened
                            cell.value = new_value
                            replacements_in_sheet += 1
                        elif new_value == "" and original_value != new_value and (text_to_find in original_value or (not case_sensitive and text_to_find.lower() in original_value.lower())):
                            # Handle case where the only content was text_to_find, resulting in empty new_value
                            cell.value = new_value
                            replacements_in_sheet += 1

            if replacements_in_sheet > 0:
                print(
                    f"    Replaced {replacements_in_sheet} occurrence(s) in sheet '{sheet.title}'.")
            total_replacements_in_workbook += replacements_in_sheet

        if total_replacements_in_workbook > 0:
            # Save the changes to the original file
            workbook.save(filepath)
            print(
                f"Workbook saved. Total replacements made: {total_replacements_in_workbook}")
        else:
            print("No occurrences of the text were found to replace in the workbook.")

    except Exception as e:
        print(f"An error occurred: {e}")
        print(
            "If you created a backup, the original file might be safe in the backup copy.")


# --- How to use the function ---
if __name__ == "__main__":
    excel_files: list[str] = [
        "circuits.xlsx",
        "constructors.xlsx",
        "drivers.xlsx",
        "pit_stops.xlsx",
        "races.xlsx",
        "results.xlsx",
        "status.xlsx",
        "weather.xlsx"
    ]

    base_path = r"C:\Users\Jantar\uni-data-warehouse\project\stage02\data\\"

    # Example usage:
    for file_name in excel_files:
        # <--- CHANGE THIS TO YOUR FILE PATH
        excel_file_path = base_path + file_name
        text_to_search_for = "/N"
        text_to_replace_it_with = ""
        # Set to True if you only want to replace exact "/N" and not "/n"
        be_case_sensitive = False

        # --- Run the replacement ---
        if excel_file_path == "your_excel_file.xlsx":
            print(
                "Please update 'excel_file_path' in the script with the actual path to your Excel file.")
        else:
            replace_text_in_excel(excel_file_path,
                                  text_to_search_for,
                                  text_to_replace_it_with,
                                  case_sensitive=be_case_sensitive,
                                  backup=True)
