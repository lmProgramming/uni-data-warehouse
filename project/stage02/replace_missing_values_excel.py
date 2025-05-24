from typing import Any
import openpyxl
import os
import shutil
from openpyxl.workbook.workbook import _WorksheetOrChartsheetLike


def replace_text_in_excel(filepath, text_to_find, text_to_replace_with, backup=True) -> Any:
    """
    Replaces text in all string cells across all sheets of an Excel file.

    Args:
        filepath (str): The path to the Excel file.
        text_to_find (str): The text string to search for.
        text_to_replace_with (str): The text string to replace with.
        case_sensitive (bool): If True (default), search is case-sensitive.
                               If False, replacement is case-insensitive specific to "/N" and "/n"
                               if text_to_find is "/N" (or its variants).
        backup (bool): If True, creates a backup of the original file.
    """
    if not os.path.exists(filepath):
        print(f"Error: File not found at {filepath}")
        return

    if backup:
        base, ext = os.path.splitext(filepath)
        backup_filepath: str = f"{base}_backup{ext}"
        try:
            shutil.copyfile(filepath, backup_filepath)
            print(f"Backup created: {backup_filepath}")
        except Exception as e:
            print(f"Could not create backup: {e}")

    try:
        workbook: openpyxl.Workbook = openpyxl.load_workbook(filepath)
        print(f"Processing workbook: {filepath}")
        total_replacements_in_workbook = 0

        for sheet_name in workbook.sheetnames:
            sheet: _WorksheetOrChartsheetLike = workbook[sheet_name]
            print(f"  Processing sheet: {sheet.title}...")
            replacements_in_sheet = 0

            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    if cell.value is not None and isinstance(cell.value, str):
                        original_value: str = cell.value
                        modified_value: str = original_value

                        if text_to_find in original_value:
                            modified_value = original_value.replace(
                                text_to_find, text_to_replace_with)

                        if modified_value != original_value:
                            cell.value = modified_value
                            replacements_in_sheet += 1

            if replacements_in_sheet > 0:
                print(
                    f"    Replaced {replacements_in_sheet} occurrence(s) in sheet '{sheet.title}'.")
            total_replacements_in_workbook += replacements_in_sheet

        if total_replacements_in_workbook > 0:
            workbook.save(filepath)
            print(
                f"Workbook saved. Total replacements made: {total_replacements_in_workbook}")
        else:
            print(
                f"No occurrences of '{text_to_find}' were found to replace in the workbook '{os.path.basename(filepath)}'.")

    except Exception as e:
        print(f"An error occurred while processing {filepath}: {e}")
        print(
            "If you created a backup, the original file might be safe in the backup copy.")


if __name__ == "__main__":
    excel_files: list[str] = [
        "circuits.xlsx",
        "constructors.xlsx",
        "drivers.xlsx",
        "pit_stops.xlsx",
        "races.xlsx",
        "results.xlsx",
        "status.xlsx",
        "weather.xlsx"
    ]

    base_path = r"C:\Users\Jantar\uni-data-warehouse\project\stage02\data\\"

    text_to_search_for = "\\N"
    text_to_replace_it_with = ""

    for file_name in excel_files:
        excel_file_path: str = os.path.join(base_path, file_name)

        print("-" * 40)
        replace_text_in_excel(excel_file_path,
                              text_to_search_for,
                              text_to_replace_it_with,
                              backup=True)
    print("-" * 40)
    print("Processing complete.")
